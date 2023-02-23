from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl

import definition_file as definition
import xpath_file as xpath


def main(mobil_name, sn_number_parameter, imei1_number, imei2_number, end_of_life_date_parameter, iteration):
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    driver.maximize_window()

    # open url
    definition.open_url(driver)

    # click on add new mobile
    definition.function(driver,
                        'xpath',
                        xpath.add_mobile,
                        'click',
                        '     Click on add new mobile....................OK',
                        'Please help me and click on add new mobile button, then press Enter. 🤗')

    # write mobile name
    definition.function(driver,
                        'xpath',
                        xpath.name,
                        'send_keys',
                        '     Write mobile name..........................OK',
                        'Please help me and write mobile name, then press Enter. 🤗',
                        mobil_name)

    # write SN
    definition.function(driver,
                        'xpath',
                        xpath.sn,
                        'send_keys',
                        '     Write mobile SN............................OK',
                        'Please help me and write mobile SN, then press Enter. 🤗',
                        sn_number_parameter)

    # write asset manager
    definition.function_with_wait(
                        driver,
                        'xpath',
                        xpath.input_assets_manager,
                        xpath.input_assets_manager_accept,
                        'send_keys',
                        '     Write asset manager........................OK',
                        'Please help me and write asset manager, then press Enter. 🤗',
                        'Zolna, Matus | 2394-Active (Matus.Zolna@magna.com)')

# write asset owner
    definition.function_with_wait(
                        driver,
                        'xpath',
                        xpath.input_assets_owner,
                        xpath.input_assets_owner_accept,
                        'send_keys',
                        '     Write asset owner..........................OK',
                        'Please help me and write asset owner, then press Enter. 🤗',
                        'Chovanec, Dominik | 2394-Active (Dominik.Chovanec@magna.com)')

# write stock keeping unit
    definition.function_with_wait(
                        driver,
                        'xpath',
                        xpath.input_stock_keeping_unit,
                        xpath.input_stock_keeping_unit_accept,
                        'send_keys',
                        '     Write stock keeping unit...................OK',
                        'Please help me and write stock keeping unit, then press Enter. 🤗',
                        'SKU25606 - Smart Phone - Apple, iPhone SE 3rd Gen')

    # select computer role
    definition.function(driver,
                        'select',
                        xpath.computer_role,
                        'select_by_index',
                        '     Pick computer role.........................OK',
                        'Please help me and pick computer role, then press Enter. 🤗',
                        '',
                        2)

    # click on save button
    definition.function(driver,
                        'xpath',
                        xpath.save_button,
                        'click',
                        '     Click on save button.......................OK',
                        'Please help me and click on save button, then press Enter. 🤗')

    # click on change status
    definition.function(driver,
                        'xpath',
                        xpath.change_status,
                        'click',
                        '     Click on change status button..............OK',
                        'Please help me and click on change status button, then press Enter. 🤗')

    time.sleep(1)

    # select proprietary
    definition.function(driver,
                        'select',
                        xpath.pick_status,
                        'select_by_index',
                        '     Pick status................................OK',
                        'Please help me and pick status, then press Enter. 🤗',
                        '',
                        3)

    # click on change status button
    definition.function(driver,
                        'xpath',
                        xpath.change_status_button,
                        'click',
                        '     Click on change status button..............OK',
                        'Please help me and click on change status button, then press Enter. 🤗')

    time.sleep(2)

    # click on edit button
    definition.function(driver,
                        'xpath',
                        xpath.edit_button,
                        'click',
                        '     Click on edit button.......................OK',
                        'Please help me and click on edit button, then press Enter. 🤗')

    # click on do not inherit check box
    definition.function(driver,
                        'xpath',
                        xpath.do_not_inherit_check_box,
                        'click',
                        '     Click on do not inherit check box..........OK',
                        'Please help me and click on do not inherit check box, then press Enter. 🤗')

    # write inventory name
    definition.function(driver,
                        'xpath',
                        xpath.inventory_name,
                        'send_keys',
                        '     Write inventory number.....................OK',
                        'Please help me and write inventory number, then press Enter. 🤗',
                        mobil_name)
# write imei1
    definition.function(driver,
                        'xpath',
                        xpath.imei1,
                        'send_keys',
                        '     Write imei1 number.........................OK',
                        'Please help me and write imei1 number, then press Enter. 🤗',
                        imei1_number)
# write imei2
    definition.function(driver,
                        'xpath',
                        xpath.imei2,
                        'send_keys',
                        '     Write imei2 number.........................OK',
                        'Please help me and write imei2 number, then press Enter. 🤗',
                        imei2_number)

    # click on do not inherit check box
    definition.function(driver,
                        'xpath',
                        xpath.end_of_life_button,
                        'click',
                        '     Click on end of life button................OK',
                        'Please help me and click on end of life button, then press Enter. 🤗')

    # write end of life date
    definition.function(driver,
                        'xpath',
                        xpath.end_of_life_input,
                        'send_keys',
                        '     Write end of life date.....................OK',
                        'Please help me and write inventory number, then press Enter. 🤗',
                        end_of_life_date_parameter)

    # write 60 days reminder
    definition.function(driver,
                        'xpath',
                        xpath.sixty_days_input,
                        'send_keys',
                        '     Write 60 days reminder.....................OK',
                        'Please help me and write 60 days reminder, then press Enter. 🤗',
                        '60')

# write mek level 1 role
    definition.function(driver,
                        'xpath',
                        xpath.input_task_recipient_role,
                        'send_keys',
                        '     Write MEK level 1 role.....................OK',
                        'Please help me and write MEK level 1 role, then press Enter. 🤗',
                        'MEG_MEK_ServiceDesk_Level_1')

    time.sleep(1)
    # click on order
    definition.function(driver,
                        'xpath',
                        xpath.order_button,
                        'click',
                        '     Click on order button......................OK',
                        'Please help me and click on order button, then press Enter. 🤗')

    # select proprietary
    definition.function(driver,
                        'select',
                        xpath.pick_proprietary,
                        'select_by_index',
                        '     Pick proprietary...........................OK',
                        'Please help me and pick proprietary, then press Enter. 🤗',
                        '',
                        1)

    # click on done
    definition.function(driver,
                        'xpath',
                        xpath.done_button,
                        'click',
                        '     Click on done button.......................OK',
                        'Please help me and click on done button, then press Enter. 🤗')

    time.sleep(2)

    # click on more
    definition.function(driver,
                        'xpath',
                        xpath.more_button,
                        'click',
                        '     Click on more button.......................OK',
                        'Please help me and click on more button, then press Enter. 🤗')

    time.sleep(1)

    # click on new window
    definition.function(driver,
                        'xpath',
                        xpath.new_window_button,
                        'click',
                        '     Click on new window button.................OK',
                        'Please help me and click on new window button, then press Enter. 🤗')

    new_window = driver.window_handles[1]
    driver.switch_to.window(new_window)

    time.sleep(4)

    definition.write_to_excel(driver.current_url, mobil_name, iteration)


# load excel with its path
path = r'inputs.xlsx'
wrkbk = openpyxl.load_workbook(path)

sh = wrkbk.active


# iterate through excel and display data
for i in range(2, sh.max_row + 1):
    mobile_name = sh.cell(row=i, column=1)
    sn_number = sh.cell(row=i, column=2)
    imei1 = sh.cell(row=i, column=3)
    imei2 = sh.cell(row=i, column=4)
    end_of_life = sh.cell(row=2, column=5)

    print(mobile_name.value)
    print(sn_number.value)
    print(imei1.value)
    print(imei2.value)
    print(" ")
    main(mobile_name.value, sn_number.value, imei1.value, imei2.value, end_of_life.value, i)

print('Process ended successfully 😊😎🤯')
